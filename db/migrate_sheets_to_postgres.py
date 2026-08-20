#!/usr/bin/env python
"""
ALSAAB AI — import the Google Sheets export into PostgreSQL.

Usage:
    # 1. Export every tab from the Google Sheet as CSV into one folder.
    #    File name must match the tab name, e.g. Partners.csv, Commissions.csv

    python db/migrate_sheets_to_postgres.py --dir ./sheets_export --dry-run
    python db/migrate_sheets_to_postgres.py --dir ./sheets_export

    # Reconciliation only, after importing:
    python db/migrate_sheets_to_postgres.py --dir ./sheets_export --verify

--dry-run parses, cleans and counts everything WITHOUT touching the database,
and prints exactly what would be inserted or skipped. Run it first.

---------------------------------------------------------------------
Data problems this handles, found in the real export
---------------------------------------------------------------------
1. The Leads tab header says "Data" where the code writes "Date", and
   "Client ID"/"Channel"/"Date" appear a second time at the end. The real
   timestamp lives in the LAST column; the first column is empty. A naive
   import leaves every created_at NULL.
2. Dates are DD/MM/YYYY HH:MM:SS. Read as MM/DD they silently corrupt.
3. Duplicate rows: the same phone was captured 4-6 times for several people,
   plus internal test rows (the 'الصعب' number, and lang_en_/lang_ar_ sessions).
4. Truncated phone numbers (9713931114, 97505157370) sit next to their
   correct versions.
"""

import argparse
import csv
import os
import re
import sys
from collections import Counter
from datetime import datetime, timezone

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "backend"))

TEST_PHONES = {"971523288001"}
TEST_SESSION_PREFIXES = ("lang_en_", "lang_ar_", "test_")


# =====================================================================
# Value cleaning
# =====================================================================

def clean_text(value):
    return str(value if value is not None else "").strip()


def parse_sheet_date(value):
    """
    Google Sheets exports as DD/MM/YYYY HH:MM:SS. Try that first, on purpose:
    01/05/2026 is 1 May, not 5 January.
    """
    text = clean_text(value)

    if not text:
        return None

    formats = (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
        "%m/%d/%Y %H:%M:%S",
    )

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None


def parse_money(value):
    text = clean_text(value)

    if not text:
        return None

    cleaned = re.sub(r"[^\d.\-]", "", text.replace(",", ""))

    try:
        return float(cleaned) if cleaned else None
    except ValueError:
        return None


def parse_int(value, default=0):
    amount = parse_money(value)
    return int(amount) if amount is not None else default


def parse_bool(value):
    return clean_text(value).lower() in ("yes", "true", "1", "on", "enabled")


def normalize_phone(value):
    """Digits only. Leading + and separators are dropped so duplicates collapse."""
    digits = re.sub(r"\D", "", clean_text(value))
    return digits or None


def normalize_partner_id(value):
    from sheet_compat import normalize_partner_id as real
    return real(value) or None


# =====================================================================
# CSV reading
# =====================================================================

# =====================================================================
# Header repair
# =====================================================================
#
# The live sheets have drifted from what the Apps Script writes. Whenever
# ensureHeaders() failed to find a header it APPENDED the correct one at the
# far right instead of fixing the broken one, so several fields now exist
# twice: a dead misspelled/split column on the left, and the real one on the
# right. Verified against the export:
#
#   Partners       col4  'Sponser Partner ID'   1/15 filled  -> dead
#                  col18 'Sponsor Partner ID'  15/15 filled  -> real
#                  col14 'Active' + col15 'Direct Customers' (a split header)
#                        0/15 filled -> dead
#                  col19 'Active Direct Customers' 15/15     -> real
#   Subscriptions  col7  'Subscription Satus'   0/15         -> dead
#                  col13 'Subscription Status' 15/15         -> real
#                  col9  'Stripe subscription ID' 0/15       -> dead
#                  col14 'Stripe Subscription ID' 15/15      -> real
#   MLMLevels      col5 'Required Course' + col6 'Workshop'  -> dead split
#                  col9  'Required Course / Workshop' 15/15  -> real
#   Referrals      col8  'Paument Status'                    -> dead
#
# Mapping the typos onto the canonical name turns each pair into a plain
# duplicate, and the first-non-empty rule below then picks the populated one
# automatically — no per-sheet special casing.

HEADER_ALIASES = {
    "sponser partner id": "Sponsor Partner ID",
    "paument status": "Payment Status",
    "subscription satus": "Subscription Status",
    "stripe subscription id": "Stripe Subscription ID",
    "required course / workshop": "Required Course / Workshop",
    "data": "Date",
}


def normalize_header(value):
    """Trim, collapse inner whitespace, then fold known typos onto the real name."""
    cleaned = re.sub(r"\s+", " ", clean_text(value))

    if not cleaned:
        return ""

    return HEADER_ALIASES.get(cleaned.lower(), cleaned)


def read_xlsx(path):
    """
    Read every tab of the Google Sheets .xlsx export.

    One File > Download > Microsoft Excel gives all 27 tabs at once, which
    beats exporting 27 separate CSVs by hand.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True)
    sheets = {}
    warnings = []

    for tab_name in workbook.sheetnames:
        worksheet = workbook[tab_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        header = [normalize_header(cell) for cell in header_row]

        duplicates = [name for name, count in Counter(h for h in header if h).items() if count > 1]
        if duplicates:
            warnings.append(
                f"{tab_name}: {duplicates} appears more than once "
                f"(drifted headers) - taking the first non-empty value"
            )

        rows = []

        for raw in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(cell is not None and clean_text(cell) for cell in raw):
                continue

            record = {}

            for index, name in enumerate(header):
                if not name:
                    continue

                value = raw[index] if index < len(raw) else None

                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")

                value = clean_text(value)

                if name not in record or not record[name]:
                    record[name] = value

            rows.append(record)

        sheets[tab_name] = rows

    workbook.close()
    return sheets, warnings


def read_csv(path):
    """
    Returns (rows, warnings).

    Duplicate headers are the reason this cannot just use DictReader: for
    'Client ID' appearing twice, DictReader silently keeps the LAST one, which
    in the Leads export is a different column than the code expects.
    Here every header keeps all its values and we take the first non-empty.
    """
    warnings = []

    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)

        try:
            header = next(reader)
        except StopIteration:
            return [], [f"{os.path.basename(path)}: file is empty"]

        header = [normalize_header(h) for h in header]

        duplicates = [name for name, count in Counter(header).items() if name and count > 1]
        if duplicates:
            warnings.append(
                f"{os.path.basename(path)}: duplicate columns {duplicates} "
                f"- taking the first non-empty value of each"
            )

        rows = []

        for raw in reader:
            if not any(clean_text(cell) for cell in raw):
                continue

            record = {}

            for index, name in enumerate(header):
                if not name:
                    continue

                value = clean_text(raw[index]) if index < len(raw) else ""

                # First non-empty wins across duplicated headers.
                if name not in record or not record[name]:
                    record[name] = value

            rows.append(record)

    return rows, warnings


def get(row, *names, default=""):
    for name in names:
        value = clean_text(row.get(name))
        if value:
            return value
    return default


# =====================================================================
# Per-sheet transforms  (sheet name -> table, columns, row builder)
# =====================================================================

def _leads(row):
    return {
        "session_id": get(row, "Session ID"),
        "client_id": get(row, "Client ID"),
        "source_partner_id": normalize_partner_id(get(row, "Source Partner ID", "Partner ID")),
        "name": get(row, "Name"),
        "phone": normalize_phone(get(row, "Phone")),
        "user_type": get(row, "User Type"),
        "business_name": get(row, "Business Name"),
        "business_type": get(row, "Business Type"),
        "pain_point": get(row, "Pain Point"),
        "channel": get(row, "Channel") or "website",
        "status": get(row, "Status") or "new",
        "email": get(row, "Email"),
        "country": get(row, "Country"),
        "created_at": parse_sheet_date(get(row, "Date")),
    }


def _partners(row):
    partner_id = normalize_partner_id(get(row, "Partner ID"))

    if not partner_id:
        return None

    return {
        "partner_id": partner_id,
        "client_id": get(row, "Client ID"),
        "sponsor_partner_id": normalize_partner_id(get(row, "Sponsor Partner ID", "Invited By")),
        "parent_partner_id": normalize_partner_id(get(row, "Parent Partner ID", "Sponsor Partner ID")),
        "partner_name": get(row, "Partner Name", "Name"),
        "phone": normalize_phone(get(row, "Phone")),
        "email": get(row, "Email") or None,
        "country": get(row, "Country"),
        "partner_rank": get(row, "Partner Rank", "Level") or "Level 1",
        "status": get(row, "Status") or "active",
        "referral_link": get(row, "Referral Link"),
        "invited_by": normalize_partner_id(get(row, "Invited By")),
        "active_direct_customers": parse_int(get(row, "Active Direct Customers")),
        "active_network_customers": parse_int(get(row, "Active Network Customers")),
        "notes": get(row, "Notes"),
        "created_at": parse_sheet_date(get(row, "Date")),
    }


def _partner_tree(row):
    ancestor = normalize_partner_id(get(row, "Ancestor Partner ID"))
    descendant = normalize_partner_id(get(row, "Descendant Partner ID"))
    depth = parse_int(get(row, "Depth"), default=0)

    # The live sheet carries one depth-0 self row per partner
    # (ALS-P00001 -> ALS-P00001). Those are not tree relations: the schema
    # constrains depth to 1..5 and the commission walk only reads 1..5.
    # Dropping them is correct, not data loss.
    if not ancestor or not descendant or not 1 <= depth <= 5:
        return None

    return {
        "ancestor_partner_id": ancestor,
        "descendant_partner_id": descendant,
        "depth": depth,
        "line_owner_partner_id": normalize_partner_id(get(row, "Line Owner Partner ID")),
        "notes": get(row, "Notes"),
        "created_at": parse_sheet_date(get(row, "Date")),
    }


def _referrals(row):
    return {
        "source_partner_id": normalize_partner_id(get(row, "Source Partner ID", "Partner ID")),
        "referral_name": get(row, "Referral Name", "Name"),
        "referral_phone": normalize_phone(get(row, "Referral Phone", "Phone")),
        "referral_email": get(row, "Referral Email", "Email"),
        "source": get(row, "Source") or "website",
        "package": get(row, "Package", "Plan Name"),
        "payment_status": get(row, "Payment Status") or "pending",
        "subscription_status": get(row, "Subscription Status") or "pending",
        "session_id": get(row, "Session ID"),
        "client_id": get(row, "Client ID"),
        "stripe_subscription_id": get(row, "Stripe Subscription ID"),
        "notes": get(row, "Notes"),
        "created_at": parse_sheet_date(get(row, "Date")),
    }


def _subscriptions(row):
    return {
        "session_id": get(row, "Session ID") or None,
        "client_id": get(row, "Client ID"),
        "source_partner_id": normalize_partner_id(get(row, "Source Partner ID", "Partner ID")),
        "plan_name": get(row, "Plan Name", "Package"),
        "package_amount": parse_money(get(row, "Package Amount")),
        "subscription_status": get(row, "Subscription Status", "Status") or "inactive",
        "stripe_customer_id": get(row, "Stripe Customer ID"),
        "stripe_subscription_id": get(row, "Stripe Subscription ID") or None,
        "current_period_start": parse_sheet_date(get(row, "Current Period Start")),
        "current_period_end": parse_sheet_date(get(row, "Current Period End")),
        "notes": get(row, "Notes"),
        "created_at": parse_sheet_date(get(row, "Date")),
    }


def _commissions(row):
    commission_id = get(row, "Commission ID")
    unique_key = get(row, "Commission Unique Key")

    if not commission_id:
        return None

    if not unique_key:
        # Rebuild it with the exact same algorithm the Apps Script used, so
        # rows written before the key column existed still de-duplicate.
        from sheet_compat import build_commission_unique_key
        unique_key = build_commission_unique_key({
            "stripe_subscription_id": get(row, "Stripe Subscription ID"),
            "payer_client_id": get(row, "Payer Client ID"),
            "source_partner_id": get(row, "Source Partner ID"),
            "beneficiary_partner_id": get(row, "Beneficiary Partner ID"),
            "commission_depth": get(row, "Commission Depth"),
            "package": get(row, "Package"),
            "period_start": get(row, "Period Start"),
        })

    depth = parse_int(get(row, "Commission Depth"), default=1)

    return {
        "commission_id": commission_id,
        "invoice_id": get(row, "Invoice ID"),
        "stripe_subscription_id": get(row, "Stripe Subscription ID"),
        "payer_client_id": get(row, "Payer Client ID"),
        "payer_name": get(row, "Payer Name"),
        "source_partner_id": normalize_partner_id(get(row, "Source Partner ID")),
        "beneficiary_partner_id": normalize_partner_id(get(row, "Beneficiary Partner ID")),
        "commission_depth": min(max(depth, 1), 5),
        "line_owner_partner_id": normalize_partner_id(get(row, "Line Owner Partner ID")),
        "partner_rank": get(row, "Partner Rank"),
        "package": get(row, "Package"),
        "package_amount": parse_money(get(row, "Package Amount")),
        "commission_percent": parse_money(get(row, "Commission %")),
        "commission_amount": parse_money(get(row, "Commission Amount")),
        "period_start": parse_sheet_date(get(row, "Period Start")),
        "period_end": parse_sheet_date(get(row, "Period End")),
        "status": (get(row, "Status") or "pending").lower(),
        "paid_date": parse_sheet_date(get(row, "Paid Date")),
        "notes": get(row, "Notes"),
        "commission_unique_key": unique_key,
        "created_at": parse_sheet_date(get(row, "Date")),
    }


def _mlm_levels(row):
    partner_id = normalize_partner_id(get(row, "Partner ID"))

    if not partner_id:
        return None

    rank = get(row, "Partner Rank") or "Level 1"
    rank_match = re.search(r"([1-5])", rank)

    return {
        "partner_id": partner_id,
        "partner_rank": rank,
        "current_level": int(rank_match.group(1)) if rank_match else 0,
        "required_sales": parse_int(get(row, "Required Sales"), default=1),
        "completed_sales": parse_int(get(row, "Completed Sales")),
        "required_course_workshop": get(row, "Required Course / Workshop", "Required Course"),
        "level_status": get(row, "Level Status") or "active",
        "next_rank": get(row, "Next Rank", "Next Level"),
        "current_package": get(row, "Current Package"),
        "subscription_status": get(row, "Subscription Status"),
        "commission_eligible": parse_bool(get(row, "Commission Eligible")),
    }


def _course_purchases(row):
    return {
        "partner_id": normalize_partner_id(get(row, "Partner ID")),
        "client_id": get(row, "Client ID"),
        "course_code": get(row, "Course Code"),
        "course_name": get(row, "Course Name"),
        "amount": parse_money(get(row, "Amount")),
        "currency": get(row, "Currency") or "USD",
        "status": get(row, "Status") or "paid",
        "stripe_payment_id": get(row, "Stripe Payment ID"),
        "stripe_customer_id": get(row, "Stripe Customer ID"),
        "notes": get(row, "Notes"),
        "paid_at": parse_sheet_date(get(row, "Paid At", "Date")),
        "refunded_at": parse_sheet_date(get(row, "Refunded At")),
    }


def _client_profiles(row):
    return {
        "session_id": get(row, "Session ID") or None,
        "client_id": get(row, "Client ID"),
        "business_name": get(row, "Business Name"),
        "business_type": get(row, "Business Type"),
        "general_description": get(row, "General Description"),
        "products": get(row, "Products"),
        "prices": get(row, "Prices"),
        "offers": get(row, "Offers"),
        "ordering": get(row, "Ordering"),
        "whatsapp": get(row, "WhatsApp"),
        "areas": get(row, "Areas"),
        "faqs": get(row, "FAQs"),
        "objections": get(row, "Objections"),
        "tone": get(row, "Tone"),
        "created_at": parse_sheet_date(get(row, "Date")),
    }


def _smart_link_events(row):
    event_id = get(row, "Event ID")

    if not event_id:
        return None

    return {
        "event_id": event_id,
        "smart_ref": get(row, "Smart Ref"),
        "client_id": get(row, "Client ID"),
        "partner_id": normalize_partner_id(get(row, "Partner ID")),
        "event_type": get(row, "Event Type"),
        "source": get(row, "Source"),
        "session_id": get(row, "Session ID"),
        "page_url": get(row, "Page URL"),
        "referrer_url": get(row, "Referrer URL"),
        "message": get(row, "Message"),
        "user_agent": get(row, "User Agent"),
        "created_at": parse_sheet_date(get(row, "Created At", "Date")),
    }


def _audit_logs(row):
    audit_id = get(row, "Audit ID")

    if not audit_id:
        return None

    def as_json(value):
        text = clean_text(value)
        if not text:
            return None
        try:
            import json
            json.loads(text)
            return text
        except ValueError:
            import json
            return json.dumps({"raw": text}, ensure_ascii=False)

    return {
        "audit_id": audit_id,
        "actor": get(row, "Actor"),
        "action": get(row, "Action"),
        "target_type": get(row, "Target Type"),
        "target_id": get(row, "Target ID"),
        "partner_id": normalize_partner_id(get(row, "Partner ID")),
        "before_json": as_json(get(row, "Before JSON")),
        "after_json": as_json(get(row, "After JSON")),
        "reason": get(row, "Reason"),
        "source": get(row, "Source"),
        "status": get(row, "Status"),
        "notes": get(row, "Notes"),
        "created_at": parse_sheet_date(get(row, "Date")),
    }


def _payout_history(row):
    payout_id = get(row, "Payout ID")

    if not payout_id:
        return None

    return {
        "payout_id": payout_id,
        "partner_id": normalize_partner_id(get(row, "Partner ID")),
        "partner_name": get(row, "Partner Name"),
        "commission_count": parse_int(get(row, "Commission Count")),
        "commission_ids": get(row, "Commission IDs"),
        "total_amount": parse_money(get(row, "Total Amount")) or 0,
        "currency": get(row, "Currency") or "AED",
        "payment_method": get(row, "Payment Method"),
        "status": (get(row, "Status") or "paid").lower(),
        "paid_date": parse_sheet_date(get(row, "Paid Date")),
        "actor": get(row, "Actor"),
        "reason": get(row, "Reason"),
        "source": get(row, "Source"),
        "notes": get(row, "Notes"),
        "created_at": parse_sheet_date(get(row, "Date")),
    }


# sheet name -> (table, ON CONFLICT target, dedupe key, transform)
#
# The conflict target is written out in full because subscriptions is backed
# by a PARTIAL unique index:
#     CREATE UNIQUE INDEX ... ON subscriptions (stripe_subscription_id)
#         WHERE stripe_subscription_id IS NOT NULL AND stripe_subscription_id <> ''
# PostgreSQL will not match a partial index from a bare "ON CONFLICT (col)" —
# the index predicate has to be repeated, otherwise it raises
# InvalidColumnReference.
#
# dedupe key: the Apps Script APPENDS a row on every save rather than updating
# in place, so a sheet can hold several versions of the same record. Keeping
# the LAST occurrence keeps the most recent version.
SHEETS = [
    ("Partners",        "partners",         "(partner_id)",            "partner_id", _partners),
    ("PartnerTree",     "partner_tree",
        "(ancestor_partner_id, descendant_partner_id, depth)",         None,         _partner_tree),
    ("MLMLevels",       "partner_levels",   "(partner_id)",            "partner_id", _mlm_levels),
    ("Leads",           "leads",            None,                      None,         _leads),
    ("Referrals",       "referrals",        None,                      None,         _referrals),
    ("ClientProfiles",  "client_profiles",  "(session_id)",            "session_id", _client_profiles),
    ("Subscriptions",   "subscriptions",
        "(stripe_subscription_id) WHERE stripe_subscription_id IS NOT NULL "
        "AND stripe_subscription_id <> ''",                            "session_id", _subscriptions),
    ("Commissions",     "commissions",      "(commission_unique_key)", "commission_unique_key", _commissions),
    ("CoursePurchases", "course_purchases", None,                      None,         _course_purchases),
    ("SmartLinkEvents", "smart_link_events", "(event_id)",              "event_id",   _smart_link_events),
    ("AuditLogs",       "audit_logs",        "(audit_id)",              "audit_id",   _audit_logs),
    ("PayoutHistory",   "payout_history",    "(payout_id)",             "payout_id",  _payout_history),
]


# A keyless table still has a combination of fields that identifies a row in
# practice. Comparing against it lets a re-import add the new rows without
# duplicating the old ones.
KEYLESS_SIGNATURES = {
    "leads": ("session_id", "phone", "created_at"),
    "referrals": ("source_partner_id", "referral_phone", "created_at"),
    "course_purchases": ("partner_id", "course_code", "stripe_payment_id"),
}


def signature_value(value):
    """
    Render one signature field the same way on both sides of the comparison.

    A datetime read back from PostgreSQL is timezone-aware and stringifies as
    "2026-05-21 15:19:38+00:00", while the same moment parsed from the sheet is
    naive and gives "2026-05-21 15:19:38". Comparing the raw strings therefore
    never matched, every stored row looked new, and a re-import duplicated all
    57 leads and 12 referrals. Normalising to a naive second-precision stamp
    makes the two agree.
    """
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value.strftime("%Y-%m-%d %H:%M:%S")

    return clean_text(value)


def dedupe_keep_last(rows, key):
    """Collapse rows sharing `key`, keeping the last (newest) one."""
    if not key:
        return rows, 0

    seen = {}
    order = []

    for row in rows:
        value = row.get(key)

        if value in (None, ""):
            order.append(("_unique", len(order)))
            seen[("_unique", len(order) - 1)] = row
            continue

        if value not in seen:
            order.append(value)

        seen[value] = row

    return [seen[k] for k in order], len(rows) - len(order)


# =====================================================================
# Import
# =====================================================================

def build_rows(source, sheet_name, transform):
    """`source` is either a dict of {tab: rows} from the xlsx, or a folder of CSVs."""
    warnings = []

    if isinstance(source, dict):
        if sheet_name not in source:
            return None, [], [f"tab '{sheet_name}' not in workbook - skipped"]
        raw_rows = source[sheet_name]
    else:
        path = os.path.join(source, f"{sheet_name}.csv")

        if not os.path.exists(path):
            return None, [], [f"{sheet_name}.csv not found - skipped"]

        raw_rows, warnings = read_csv(path)

    rows = []
    dropped = 0

    for raw in raw_rows:
        record = transform(raw)

        if record is None:
            dropped += 1
            continue

        rows.append(record)

    if dropped:
        warnings.append(
            f"{sheet_name}: {dropped} row(s) dropped - missing a required id, "
            f"or (PartnerTree) a depth-0 self row"
        )

    return rows, raw_rows, warnings


def flag_suspicious_leads(rows):
    """Reports likely-duplicate and test rows. Does NOT delete them."""
    notes = []
    by_phone = Counter(r["phone"] for r in rows if r.get("phone"))

    repeats = {phone: n for phone, n in by_phone.items() if n > 1}
    if repeats:
        top = sorted(repeats.items(), key=lambda kv: -kv[1])[:5]
        notes.append(
            "repeated phones: "
            + ", ".join(f"{phone} x{count}" for phone, count in top)
            + (f" (+{len(repeats) - len(top)} more)" if len(repeats) > len(top) else "")
        )

    test_rows = [
        r for r in rows
        if r.get("phone") in TEST_PHONES
        or str(r.get("session_id") or "").startswith(TEST_SESSION_PREFIXES)
    ]
    if test_rows:
        notes.append(f"internal/test rows: {len(test_rows)}")

    short = [r for r in rows if r.get("phone") and len(r["phone"]) < 11]
    if short:
        notes.append(f"suspiciously short phones: {len(short)} ({', '.join(r['phone'] for r in short[:3])})")

    missing_date = [r for r in rows if not r.get("created_at")]
    if missing_date:
        notes.append(f"rows with unparseable date: {len(missing_date)}")

    return notes


def insert_rows(cur, table, rows, conflict_target):
    if not rows:
        return 0

    columns = list(rows[0].keys())
    placeholders = ", ".join(["%s"] * len(columns))
    column_list = ", ".join(columns)

    conflict = f" ON CONFLICT {conflict_target} DO NOTHING" if conflict_target else ""
    sql = f"INSERT INTO {table} ({column_list}) VALUES ({placeholders}){conflict}"

    inserted = 0

    for row in rows:
        try:
            cur._cursor.execute(sql, tuple(row[c] for c in columns))
            inserted += cur._cursor.rowcount
        except Exception as error:
            print(f"    ! row failed in {table}: {type(error).__name__}: {error}")
            raise

    return inserted


def main():
    parser = argparse.ArgumentParser(description="Import the Google Sheets export into PostgreSQL")
    parser.add_argument("--xlsx", help="the .xlsx export (File > Download > Microsoft Excel)")
    parser.add_argument("--dir", help="alternative: a folder of per-tab CSV exports")
    parser.add_argument("--dry-run", action="store_true", help="parse and report, do not write")
    parser.add_argument("--verify", action="store_true", help="compare source counts to table counts")
    parser.add_argument("--force", action="store_true",
                        help="re-import keyless tables even if they already hold rows (duplicates them)")
    args = parser.parse_args()

    if not args.xlsx and not args.dir:
        print("Give either --xlsx <file> or --dir <folder>.")
        return 2

    all_warnings = []

    if args.xlsx:
        if not os.path.isfile(args.xlsx):
            print(f"Not a file: {args.xlsx}")
            return 2
        source, all_warnings = read_xlsx(args.xlsx)
        source_label = os.path.abspath(args.xlsx)
    else:
        if not os.path.isdir(args.dir):
            print(f"Not a folder: {args.dir}")
            return 2
        source = args.dir
        source_label = os.path.abspath(args.dir)

    print("=" * 66)
    print("ALSAAB AI - Google Sheets -> PostgreSQL")
    print(f"source : {source_label}")
    print(f"mode   : {'DRY RUN (nothing is written)' if args.dry_run else 'VERIFY' if args.verify else 'IMPORT'}")
    print("=" * 66)

    parsed = {}

    for sheet_name, table, conflict_target, dedupe_key, transform in SHEETS:
        rows, raw_rows, warnings = build_rows(source, sheet_name, transform)
        all_warnings.extend(warnings)

        if rows is None:
            print(f"{sheet_name:<16} -> {table:<18} (no CSV)")
            continue

        rows, collapsed = dedupe_keep_last(rows, dedupe_key)

        parsed[table] = (rows, conflict_target)
        print(f"{sheet_name:<16} -> {table:<18} {len(raw_rows):>5} rows in source, {len(rows):>5} ready")

        if collapsed:
            print(f"                    . {collapsed} older version(s) of the same "
                  f"{dedupe_key} collapsed, newest kept")

        if sheet_name == "Leads":
            for note in flag_suspicious_leads(rows):
                print(f"                    . {note}")

    if all_warnings:
        print("\nWarnings:")
        for warning in all_warnings:
            print(f"  ! {warning}")

    if args.dry_run:
        print("\nDry run complete. Nothing was written.")
        print("Re-run without --dry-run to import.")
        return 0

    from db import get_connection, USING_POSTGRES

    if not USING_POSTGRES:
        print("\nDATABASE_URL is not set - refusing to import into SQLite.")
        return 2

    conn = get_connection()

    try:
        cur = conn.cursor()

        if args.verify:
            print("\nReconciliation (CSV vs database):")
            ok = True

            for table, (rows, _) in parsed.items():
                cur.execute(f"SELECT COUNT(*) FROM {table}")
                in_db = cur.fetchone()[0]
                match = "OK" if in_db >= len(rows) else "MISMATCH"
                if in_db < len(rows):
                    ok = False
                print(f"  {table:<20} csv={len(rows):>5}  db={in_db:>5}  {match}")

            print("\n" + ("All tables reconciled." if ok else "Some tables are short - re-run the import."))
            return 0 if ok else 1

        # partners must land before anything that references a partner id.
        cur._cursor.execute("SET CONSTRAINTS ALL DEFERRED")

        print("\nImporting:")
        total = 0

        for table, (rows, conflict_target) in parsed.items():
            # leads and referrals have no unique key, so ON CONFLICT cannot
            # protect them. Instead of skipping the whole table on a re-import
            # -- which would also lock out genuinely new rows -- compare each
            # incoming row against a natural signature of what is already
            # stored, and insert only what is missing.
            if rows and not conflict_target:
                signature = KEYLESS_SIGNATURES.get(table)

                if signature and not args.force:
                    columns = ", ".join(signature)
                    cur.execute(f"SELECT {columns} FROM {table}")
                    seen = {
                        tuple(signature_value(v) for v in existing_row)
                        for existing_row in cur._cursor.fetchall()
                    }

                    fresh = [
                        row for row in rows
                        if tuple(signature_value(row.get(col)) for col in signature) not in seen
                    ]

                    already = len(rows) - len(fresh)

                    if already:
                        print(f"  {table:<20} {already} row(s) already present, skipped")

                    rows = fresh

                    if not rows:
                        print(f"  {table:<20} nothing new to add")
                        continue

                elif not args.force:
                    cur.execute(f"SELECT COUNT(*) FROM {table}")
                    existing = cur._cursor.fetchone()[0]

                    if existing:
                        print(f"  {table:<20} SKIPPED - already has {existing} rows and no unique key")
                        continue

            inserted = insert_rows(cur, table, rows, conflict_target)
            skipped = len(rows) - inserted
            total += inserted
            print(f"  {table:<20} inserted={inserted:>5}  skipped(dupe)={skipped:>5}")

        # Continue partner ids after the highest imported ALS-Pxxxxx.
        cur._cursor.execute(
            """
            SELECT setval('partner_id_seq', COALESCE((
                SELECT MAX((regexp_match(partner_id, '^ALS-P(\\d+)$'))[1]::INT)
                FROM partners WHERE partner_id ~ '^ALS-P\\d+$'
            ), 0) + 1, false)
            """
        )
        cur._cursor.execute("SELECT last_value FROM partner_id_seq")
        next_id = cur._cursor.fetchone()[0]

        conn.commit()

        print(f"\nDone. {total} rows inserted.")
        print(f"Next generated partner id will be ALS-P{str(next_id).zfill(5)}")
        flag = f'--xlsx "{args.xlsx}"' if args.xlsx else f'--dir "{args.dir}"'
        print(f"\nNow run:  python db/migrate_sheets_to_postgres.py {flag} --verify")
        return 0

    except Exception as error:
        conn.rollback()
        print(f"\nIMPORT FAILED - rolled back, database unchanged.")
        print(f"{type(error).__name__}: {error}")
        return 1

    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
